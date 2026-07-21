# ==========================================================
# NEW INCOMING CHECK IN SYSTEM
# ==========================================================

import os
import sys
import subprocess
from datetime import datetime
import tkinter as tk
from tkinter import ttk, messagebox, simpledialog, filedialog

try:
    from tkcalendar import DateEntry
    from PIL import Image, ImageTk
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image as OpenpyxlImage
    from openpyxl.drawing.spreadsheet_drawing import TwoCellAnchor, AnchorMarker
except ImportError:
    root = tk.Tk()
    root.withdraw()
    messagebox.showerror("Dependency Error", "Please ensure openpyxl, pillow, and tkcalendar are installed via pip first.")
    sys.exit()

# ==========================================================
# FILES & DEFAULTS
# ==========================================================

EXCEL_FILE = "New_Incoming_Check_In.xlsx"
LISTS_SHEET_NAME = "Lists"

DEFAULT_SUPPLIER = [
    "AVNET ASIA", "DIGIKEY ELECTRONICS", "ELEMENT 14", "FUTURE ELECTRONICS",
    "MOUSER ELECTRONICS", "GPV ASIA THAILAND", "HONG KONG YIHAO", "PHYTECH",
    "WIN SOURCE ELECTRONICS", "HONENG ELEC", "UNI BETTER", "LCSC",
    "NELSON MILLER GROUP", "E-STAR", "WURTH ELEKTRONIK"
]

DEFAULT_COMMODITY = [
    "PCB", "CAPACITOR", "RESISTOR", "DIODE", "FERRITE", "TRANSISTOR", "IC"
]

# Column J (10) = Picture, Column K (11) = Remark
HEADERS = [
    "Store Receive Date", "IQA Receive Date", "Supplier", 
    "MPN No.", "Part No.", "DC/Lot No.", "Quantity", 
    "Commodity", "Status", "Picture", "Remark"
]

selected_image_path = None  # Global variable for image path

# ==========================================================
# EXCEL HELPER FUNCTIONS
# ==========================================================

def apply_excel_formatting(ws):
    """Ensures uniform Calibri font face, row heights, and wrapped column grids."""
    thin_side = Side(style="thin", color="000000")
    cell_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    
    # Format Headers (Row 1)
    ws.row_dimensions[1].height = 25
    for col in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(name="Calibri", size=11, bold=True)
        if col in [3, 4, 5]:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        else:
            cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = cell_border
        ws.column_dimensions[get_column_letter(col)].width = 22

    # Format Data Rows (Row 2 onwards)
    data_font = Font(name="Calibri", size=11)
    for row in range(2, ws.max_row + 1):
        ws.row_dimensions[row].height = 65  # Height spacing to accommodate centered picture cleanly
        for col in range(1, len(HEADERS) + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = data_font
            cell.border = cell_border
            if col in [3, 4, 5]:
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")

def initialize_excel_database():
    """Ensures Excel file and Lists sheet exist with default values if missing."""
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        wb.remove(wb.active)
        
        ws_lists = wb.create_sheet(title=LISTS_SHEET_NAME)
        ws_lists.cell(row=1, column=1, value="Supplier").font = Font(name="Calibri", size=11, bold=True)
        ws_lists.cell(row=1, column=2, value="Commodity").font = Font(name="Calibri", size=11, bold=True)
        
        for idx, item in enumerate(DEFAULT_SUPPLIER, start=2):
            ws_lists.cell(row=idx, column=1, value=item)
            
        for idx, item in enumerate(DEFAULT_COMMODITY, start=2):
            ws_lists.cell(row=idx, column=2, value=item)

        wb.save(EXCEL_FILE)
        wb.close()

def load_lists_from_excel():
    """Reads Supplier and Commodity entries directly from the 'Lists' sheet."""
    initialize_excel_database()
    suppliers = []
    commodities = []
    
    try:
        wb = load_workbook(EXCEL_FILE, data_only=True)
        if LISTS_SHEET_NAME in wb.sheetnames:
            ws = wb[LISTS_SHEET_NAME]
            for row in range(2, ws.max_row + 1):
                val = ws.cell(row=row, column=1).value
                if val:
                    suppliers.append(str(val).strip())
            for row in range(2, ws.max_row + 1):
                val = ws.cell(row=row, column=2).value
                if val:
                    commodities.append(str(val).strip())
        wb.close()
    except Exception as e:
        print(f"Error loading lists: {e}")

    if not suppliers:
        suppliers = DEFAULT_SUPPLIER.copy()
    if not commodities:
        commodities = DEFAULT_COMMODITY.copy()

    return suppliers, commodities

# Load initial lists
supplier_list, commodity_list = load_lists_from_excel()

# ==========================================================
# MAIN WINDOW SETUP
# ==========================================================

root = tk.Tk()
root.title("New Incoming Check In")
root.geometry("800x800")
root.resizable(True, True)
root.configure(bg="white")
root.grid_columnconfigure(0, weight=1)

root.option_add("*TCombobox*Listbox.font", ("Calibri", 13))

# ==========================================================
# LOGO (AUTO-RATIO)
# ==========================================================

try:
    img = Image.open("Kaltech Logo.png")
    target_width = 120
    ratio = target_width / float(img.size[0])
    target_height = int(float(img.size[1]) * ratio)
    img = img.resize((target_width, target_height), Image.Resampling.LANCZOS)
    logo = ImageTk.PhotoImage(img)
    logo_label = tk.Label(root, image=logo, bg="white")
    logo_label.pack(pady=(10, 5))
except Exception as e:
    print(f"Logo failed to load: {e}")

# ==========================================================
# TITLE
# ==========================================================

title = tk.Label(root, text="NEW INCOMING CHECK IN", font=("Calibri", 22, "bold"), bg="white", fg="#1F4E79")
title.pack()

subtitle = tk.Label(root, text="Incoming Material Inspection System", font=("Calibri", 11, "italic"), bg="white", fg="gray")
subtitle.pack(pady=(0, 10))

# ==========================================================
# FORM FRAME
# ==========================================================

frame = tk.Frame(root, bg="white")
frame.pack(padx=20)

# 1. Store Date
tk.Label(frame, text="Store Receive Date", font=("Calibri", 13, "bold"), bg="white").grid(row=0, column=0, padx=10, pady=6, sticky="w")
store_date = DateEntry(frame, width=18, date_pattern="dd/mm/yyyy", font=("Calibri", 13))
store_date.grid(row=0, column=1, sticky="w")

# 2. IQA Date
tk.Label(frame, text="IQA Receive Date", font=("Calibri", 13, "bold"), bg="white").grid(row=1, column=0, padx=10, pady=6, sticky="w")
iqa_date = DateEntry(frame, width=18, date_pattern="dd/mm/yyyy", font=("Calibri", 13))
iqa_date.grid(row=1, column=1, sticky="w")

# 3. Supplier
tk.Label(frame, text="Supplier", font=("Calibri", 13, "bold"), bg="white").grid(row=2, column=0, padx=10, pady=6, sticky="w")
supplier = ttk.Combobox(frame, values=supplier_list, width=37, state="readonly", font=("Calibri", 13))
supplier.grid(row=2, column=1, sticky="w")

# ==========================================================
# DYNAMIC LIST FUNCTIONS
# ==========================================================

def add_supplier():
    new_supplier = simpledialog.askstring("Add New Supplier", "Enter supplier name:")
    if new_supplier:
        new_supplier = new_supplier.upper().strip()
        if new_supplier not in supplier_list:
            try:
                wb = load_workbook(EXCEL_FILE)
                ws = wb[LISTS_SHEET_NAME] if LISTS_SHEET_NAME in wb.sheetnames else wb.create_sheet(title=LISTS_SHEET_NAME)
                
                row = 2
                while ws.cell(row=row, column=1).value is not None:
                    row += 1
                ws.cell(row=row, column=1, value=new_supplier)
                
                wb.save(EXCEL_FILE)
                wb.close()

                supplier_list.append(new_supplier)
                supplier["values"] = supplier_list
                supplier.set(new_supplier)
                messagebox.showinfo("Success", "New supplier added.")
            except Exception as e:
                messagebox.showerror("Error", f"Could not update Excel lists: {e}")

def delete_supplier():
    current_val = supplier.get()
    if not current_val:
        messagebox.showwarning("Selection Empty", "Please select a supplier from the list first to delete.")
        return
    
    if messagebox.askyesno("Confirm Delete", f"Are you sure you want to delete supplier '{current_val}'?"):
        try:
            supplier_list.remove(current_val)
            supplier["values"] = supplier_list
            supplier.set("")

            wb = load_workbook(EXCEL_FILE)
            if LISTS_SHEET_NAME in wb.sheetnames:
                ws = wb[LISTS_SHEET_NAME]
                for r in range(2, ws.max_row + 1):
                    ws.cell(row=r, column=1).value = None
                for idx, item in enumerate(supplier_list, start=2):
                    ws.cell(row=idx, column=1, value=item)
            wb.save(EXCEL_FILE)
            wb.close()
            messagebox.showinfo("Deleted", f"'{current_val}' has been removed from the list.")
        except Exception as e:
            messagebox.showerror("Error", f"Could not delete supplier: {e}")

def add_commodity():
    new_commodity = simpledialog.askstring("Add New Commodity", "Enter commodity name:")
    if new_commodity:
        new_commodity = new_commodity.upper().strip()
        if new_commodity not in commodity_list:
            try:
                wb = load_workbook(EXCEL_FILE)
                ws = wb[LISTS_SHEET_NAME] if LISTS_SHEET_NAME in wb.sheetnames else wb.create_sheet(title=LISTS_SHEET_NAME)
                
                row = 2
                while ws.cell(row=row, column=2).value is not None:
                    row += 1
                ws.cell(row=row, column=2, value=new_commodity)
                
                wb.save(EXCEL_FILE)
                wb.close()

                commodity_list.append(new_commodity)
                commodity["values"] = commodity_list
                commodity.set(new_commodity)
                messagebox.showinfo("Success", "New commodity added.")
            except Exception as e:
                messagebox.showerror("Error", f"Could not update Excel lists: {e}")

def delete_commodity():
    current_val = commodity.get()
    if not current_val:
        messagebox.showwarning("Selection Empty", "Please select a commodity from the list first to delete.")
        return
    
    if messagebox.askyesno("Confirm Delete", f"Are you sure you want to delete commodity '{current_val}'?"):
        try:
            commodity_list.remove(current_val)
            commodity["values"] = commodity_list
            commodity.set("")

            wb = load_workbook(EXCEL_FILE)
            if LISTS_SHEET_NAME in wb.sheetnames:
                ws = wb[LISTS_SHEET_NAME]
                for r in range(2, ws.max_row + 1):
                    ws.cell(row=r, column=2).value = None
                for idx, item in enumerate(commodity_list, start=2):
                    ws.cell(row=idx, column=2, value=item)
            wb.save(EXCEL_FILE)
            wb.close()
            messagebox.showinfo("Deleted", f"'{current_val}' has been removed from the list.")
        except Exception as e:
            messagebox.showerror("Error", f"Could not delete commodity: {e}")

add_supplier_btn = tk.Button(frame, text="+ Add Supplier", command=add_supplier, bg="#DDDDDD", font=("Calibri", 11, "bold"))
add_supplier_btn.grid(row=2, column=2, padx=10)

def create_entry(row, text):
    tk.Label(frame, text=text, font=("Calibri", 13, "bold"), bg="white").grid(row=row, column=0, padx=10, pady=6, sticky="w")
    entry = tk.Entry(frame, width=40, font=("Calibri", 13))
    entry.grid(row=row, column=1, pady=6, sticky="w")
    return entry

mpn = create_entry(3, "MPN No.")
part_no = create_entry(4, "Part No.")
lot_no = create_entry(5, "DC/Lot No.")
quantity = create_entry(6, "Quantity")

# Commodity
tk.Label(frame, text="Commodity", font=("Calibri", 13, "bold"), bg="white").grid(row=7, column=0, padx=10, pady=6, sticky="w")
commodity = ttk.Combobox(frame, values=commodity_list, width=37, state="readonly", font=("Calibri", 13))
commodity.grid(row=7, column=1, sticky="w")

add_commodity_btn = tk.Button(frame, text="+ Add Commodity", command=add_commodity, bg="#DDDDDD", font=("Calibri", 11, "bold"))
add_commodity_btn.grid(row=7, column=2, padx=10)

# Status
tk.Label(frame, text="Status", font=("Calibri", 13, "bold"), bg="white").grid(row=8, column=0, padx=10, pady=6, sticky="w")
status = ttk.Combobox(frame, values=["ACCEPT", "REJECT", "WAIVER", "QA PASS", "ON HOLD"], width=37, state="readonly", font=("Calibri", 13))
status.grid(row=8, column=1, sticky="w")

# Remark
remark = create_entry(9, "Remark")

# Context Menus for Right Click
supplier_menu = tk.Menu(root, tearoff=0)
supplier_menu.add_command(label="❌ Delete Selected Supplier", command=delete_supplier)

commodity_menu = tk.Menu(root, tearoff=0)
commodity_menu.add_command(label="❌ Delete Selected Commodity", command=delete_commodity)

supplier.bind("<Button-3>", lambda event: supplier_menu.post(event.x_root, event.y_root))
commodity.bind("<Button-3>", lambda event: commodity_menu.post(event.x_root, event.y_root))

# ==========================================================
# PICTURE UPLOAD SECTION
# ==========================================================

tk.Label(frame, text="Picture", font=("Calibri", 13, "bold"), bg="white").grid(row=10, column=0, padx=10, pady=6, sticky="w")

img_preview_label = tk.Label(frame, text="No image selected", font=("Calibri", 10, "italic"), bg="#F0F0F0", width=30, height=3)
img_preview_label.grid(row=10, column=1, sticky="w", pady=6)

def choose_picture():
    global selected_image_path
    file_path = filedialog.askopenfilename(
        title="Select Material Picture",
        filetypes=[("Image Files", "*.png *.jpg *.jpeg *.bmp *.gif")]
    )
    if file_path:
        selected_image_path = file_path
        img = Image.open(file_path)
        img.thumbnail((120, 60))
        img_tk = ImageTk.PhotoImage(img)
        img_preview_label.config(image=img_tk, text="", width=120, height=60)
        img_preview_label.image = img_tk

upload_btn = tk.Button(frame, text="📷 Upload Picture", command=choose_picture, bg="#E1E1E1", font=("Calibri", 11, "bold"))
upload_btn.grid(row=10, column=2, padx=10)

# ==========================================================
# SAVE DATA PIPELINE (DYNAMIC MONTHS + CENTERED PICTURE)
# ==========================================================

def check_in():
    global selected_image_path
    
    if supplier.get() == "" or mpn.get() == "" or part_no.get() == "" or status.get() == "":
        messagebox.showwarning("Incomplete Data", "Please complete all required fields.")
        return

    try:
        qty = int(quantity.get())
    except ValueError:
        messagebox.showwarning("Invalid Quantity", "Quantity must be a valid number.")
        return

    try:
        date_obj = datetime.strptime(store_date.get(), "%d/%m/%Y")
        sheet_name = date_obj.strftime("%b %Y")
    except Exception:
        sheet_name = "General"

    try:
        initialize_excel_database()
        wb = load_workbook(EXCEL_FILE)

        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(title=sheet_name, index=0)
            for col, text in enumerate(HEADERS, start=1):
                ws.cell(row=1, column=col).value = text
            ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}1"
        else:
            ws = wb[sheet_name]

        row = ws.max_row + 1

        # Data mapping: Column J (10) = "", Column K (11) = Remark
        data = [
            store_date.get(), iqa_date.get(), supplier.get(),
            mpn.get().upper(), part_no.get().upper(), lot_no.get().upper(),
            qty, commodity.get(), status.get(), "", remark.get()
        ]

        for col, value in enumerate(data, start=1):
            cell = ws.cell(row=row, column=col)
            cell.value = value

        colors = {
            "ACCEPT": "009900", "REJECT": "FF0000", "WAIVER": "00B0F0",
            "QA PASS": "66FF33", "ON HOLD": "FFFF00"
        }
        if status.get() in colors:
            ws.cell(row=row, column=9).fill = PatternFill("solid", fgColor=colors[status.get()])

        # Center-Aligned Image Insertion via TwoCellAnchor
        if selected_image_path and os.path.exists(selected_image_path):
            try:
                img_to_insert = OpenpyxlImage(selected_image_path)

                col_idx = 9      # Column J (0-indexed)
                row_idx = row - 1 # Current Row (0-indexed)

                cell_w_px = 154  # Width ~22 chars = ~154px
                cell_h_px = 86   # Height 65pt = ~86px

                # Scale proportionally within cell bounds
                max_w, max_h = 135, 72
                w, h = img_to_insert.width, img_to_insert.height
                ratio = min(max_w / float(w), max_h / float(h))
                new_w = int(w * ratio)
                new_h = int(h * ratio)

                EMU_PER_PX = 9525
                offset_x_emu = int((cell_w_px - new_w) / 2) * EMU_PER_PX
                offset_y_emu = int((cell_h_px - new_h) / 2) * EMU_PER_PX

                marker_from = AnchorMarker(
                    col=col_idx, 
                    colOff=offset_x_emu, 
                    row=row_idx, 
                    rowOff=offset_y_emu
                )

                marker_to = AnchorMarker(
                    col=col_idx, 
                    colOff=offset_x_emu + (new_w * EMU_PER_PX), 
                    row=row_idx, 
                    rowOff=offset_y_emu + (new_h * EMU_PER_PX)
                )

                img_to_insert.anchor = TwoCellAnchor(_from=marker_from, to=marker_to, editAs="oneCell")
                ws.add_image(img_to_insert)

            except Exception as img_err:
                print(f"Could not embed picture: {img_err}")

        apply_excel_formatting(ws)

        wb.save(EXCEL_FILE)
        wb.close()
        
        messagebox.showinfo("Success", f"Material checked in successfully under tab [{sheet_name}].")
        clear_form()
        
    except Exception as e:
        messagebox.showerror("File Write Error", f"Make sure '{EXCEL_FILE}' is closed first.\nDetails: {e}")

# ==========================================================
# CLEAR FORM
# ==========================================================

def clear_form():
    global selected_image_path
    mpn.delete(0, tk.END)
    part_no.delete(0, tk.END)
    lot_no.delete(0, tk.END)
    quantity.delete(0, tk.END)
    remark.delete(0, tk.END)
    supplier.set("")
    commodity.set("")
    status.set("")
    selected_image_path = None
    img_preview_label.config(image="", text="No image selected", width=30, height=3)
    mpn.focus()

# ==========================================================
# OPEN EXCEL
# ==========================================================

def open_excel():
    try:
        os.startfile(EXCEL_FILE)
    except Exception:
        subprocess.Popen(["start", EXCEL_FILE], shell=True)

# ==========================================================
# BUTTONS
# ==========================================================

button_frame = tk.Frame(root, bg="white")
button_frame.pack(pady=15)

check_btn = tk.Button(button_frame, text="CHECK IN", font=("Calibri", 13, "bold"), bg="#009900", fg="white", width=16, height=2, command=check_in)
check_btn.grid(row=0, column=0, padx=20)

excel_btn = tk.Button(button_frame, text="OPEN EXCEL", font=("Calibri", 13, "bold"), bg="#0066CC", fg="white", width=16, height=2, command=open_excel)
excel_btn.grid(row=0, column=1, padx=20)

# ==========================================================
# FOCUS EVENT MAPPINGS
# ==========================================================

def focus_next(event):
    event.widget.tk_focusNext().focus()
    return "break"

for widget in [mpn, part_no, lot_no, quantity, remark]:
    widget.bind("<Return>", focus_next)

supplier.bind("<<ComboboxSelected>>", lambda e: mpn.focus())
commodity.bind("<<ComboboxSelected>>", lambda e: status.focus())
status.bind("<<ComboboxSelected>>", lambda e: remark.focus())

root.bind("<F12>", lambda e: check_in())

mpn.focus()
root.mainloop()